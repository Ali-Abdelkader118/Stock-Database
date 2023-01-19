import sqlite3
import openpyxl
from barcode import Code39
from barcode.writer import ImageWriter
import arabic_reshaper
from bidi.algorithm import get_display


#Creating The Database File And opening It
db = sqlite3.connect("Stock.db")
c = db.cursor()


def search_data():
    #Getting The User Input And Making It UPPER To Match The Database
    code = input('Input Your Code: ').upper()
    #Checking If The Code Is Correct
    if len(code) != 6 :
        print("Code Incorrect")
    else:
        #Searching For The Code In The Database
        d = c.execute("SELECT * FROM Main WHERE Barcode=?",(code,),).fetchall()
        print(d)




def create_barcode():
    #Getting Data From The Database
    c.execute('''select Barcode , Product_Name from Main''')
    data = c.fetchall()
    #Looping For Creating The Barcode
    for i in data:
        print(r"ID : ", i[0])
        print(r"Name :", i[1])
        print("\n")
        #Creating The Barcode
        my_code = Code39(i[0], writer=ImageWriter())
        #Saving The Barcode As Image
        my_code.save(f"Barcodes/{i[0]}")



def excel_data():
    #Getting The User Input For The File Name And Number Of Rows 
    n_rows = int(input("Enter Number Of Rows: "))
    #Opening the Excel
    wb = openpyxl.load_workbook("Data/Main.xlsx")
    ws = wb.active
    ID = 1
    #Creating The Table If Not Exist
    c.execute("DROP TABLE IF EXISTS Main")
    c.execute("CREATE TABLE Main(Barcode TEXT NOT NULL UNIQUE, Product_Name TEXT NOT NULL , QTY TEXT NOT NULL)")
    while ID <= n_rows:
        # Getting The Data From Excel
        Barcode = ws[f"A{ID}"].value
        Product_Name = ws[f"B{ID}"].value
        n = 1
        #Inserting The Data Into Sqlite Database
        c.execute("INSERT INTO Main(Barcode , Product_Name, QTY) values(?, ?, ?)",(Barcode , Product_Name, n))
        #Printing The Final Result
        print("Data Inserted in the table: ")
        data=c.execute('''SELECT * FROM Main''')
        for row in data:
            print(row)
        #increasing the ID Count to Switch to next Row
        ID = ID + 1
    db.commit()



def add_item():
    #Getting The User Input
    Item_ID = input("Enter Item ID: ").upper()
    n = int(input("Enter Added Number: "))
    #Checking If The Product-Code Exist
    if len(Item_ID) > 6 or len(Item_ID) < 6 and type(Item_ID) != str():
        print("Code Incorrect")
    else:
        #Searching For The Product
        d = c.execute("SELECT QTY FROM Main WHERE Barcode=?",(Item_ID,),).fetchone()
        #Looping To Convert The Data From Tuple To Int
        for i in d :
            #Converting The Type Of i From Tuple To Int
            i = int(i)
            print(f"Old Number : {i}")
            #Adding The New Given Number To The Old Number
            new_n = i + n
            print(f"New Number : {new_n}")
            #Updating The Database With The New Data
            c.execute(f"UPDATE {Item_ID[0]} SET QTY={new_n} WHERE Barcode=?",(Item_ID,),)
        #printing The Results
        d = c.execute("SELECT * FROM Main WHERE Barcode=?",(Item_ID,),).fetchall()
        print(d)
        db.commit()



def remove_item():
    #Getting The User Input
    Item_ID = input("Enter Item ID: ").upper()
    n = int(input("Enter Removed Number: "))
    #Checking If The Product-Code Exist
    if len(Item_ID) > 6 or len(Item_ID) < 6 and type(Item_ID) != str():
        print("Code Incorrect")
    else:
        #Searching For The Product
        d = c.execute("SELECT QTY FROM Main WHERE Barcode=?",(Item_ID,),).fetchone()
        #Looping To Convert The Data From Tuple To Int
        for i in d :
            #Converting The Type Of i From Tuple To Int
            i = int(i)
            print(f"Old Number : {i}")
            #Removing The New Given Number From The Old Number
            new_n = i - n
            print(f"New Number : {new_n}")
            #Updating The Database With The New Data
            c.execute(f"UPDATE {Item_ID[0]} SET QTY={new_n} WHERE Barcode=?",(Item_ID,),)
        #printing The Results
        d = c.execute("SELECT * FROM Main WHERE Barcode=?",(Item_ID,),).fetchall()
        print(d)
        db.commit()


def search_name():
    #Getting The User Input
    code = input('Input Your Product Name: ')
    #Searching For The Name In The Database (or Anything Like It)
    d = c.execute("SELECT * FROM Main WHERE Product_Name LIKE ?",("%" + code + "%",),).fetchall()
    #Looping To Convert Type Of Data From Tuple To List
    for i in d :
        print("Product Code:" + i[0])
        #Correcting The Arabic Text RTL And Joining The Letters
        text = arabic_reshaper.reshape(i[1])
        print("Product Name:" + get_display(text))
        print("Product QTY:" + i[2])
        print("\n")



f = input("Hi , What Would You Like To Do ? ( Search - ADD - Remove - Create Barcodes) : ").lower()
if f != "search" and f != 'add' and f != 'remove' and f != r"create barcodes" and f != "excel":
    print("Invalid Command")
elif f == 'search' :
    search_name()
elif f == 'add' :
    add_item()
elif f == 'remove' :
    remove_item()
elif f ==  r"create barcodes" :
    create_barcode()
elif f == r"excel" :
    excel_data()